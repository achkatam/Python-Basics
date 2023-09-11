import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    # returns an workbook object
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Let's find out the maximum number of rows in the sheet
    # print(sheet.max_row)
    last_row = sheet.max_row

    for row in range(2, last_row + 1):
        cell = sheet.cell(row, 3)
        # print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    cell_to_rename = sheet.cell(1, 4)
    cell_to_rename.value = 'corrected price'

    values = Reference(sheet,
                       min_row=2,
                       max_row=last_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save(filename)


if __name__ == '__main__':
    process_workbook('transactions.xlsx')